# -*- coding: utf8 -*-
import datetime
import requests
import openpyxl
import gitlab
import common
import configparser, os
#from datetime import datetime

GITLAB_URL = 'http://192.168.0.6'
GITLAB_ACESS_TOKEN = "msFPgRC4QrCt8b8Fb5vm"

EXCEL_REP_INDEX = 1
EXCEL_BRANCH_INDEX = 2
EXCEL_COMMIT_INDEX = 3
EXCEL_DATE_INDEX = 4
EXCEL_USER_INDEX = 5
EXCEL_CID = 6

def create_excel():
    wb = openpyxl.Workbook() 
    sheet = wb.active 
    sheet.title ='BK72XX'
    sheet.cell(row=1, column=EXCEL_REP_INDEX, value="仓库") #写入单元格
    sheet.cell(row=1, column=EXCEL_BRANCH_INDEX, value="分支")
    sheet.cell(row=1, column=EXCEL_COMMIT_INDEX, value="commit")
    sheet.cell(row=1, column=EXCEL_DATE_INDEX, value="提交日期")
    sheet.cell(row=1, column=EXCEL_USER_INDEX, value="提交人")
    sheet.cell(row=1, column=EXCEL_CID, value="SHA")
    #set width
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 70
    sheet.column_dimensions["D"].width = 20
    sheet.column_dimensions["E"].width = 5
    sheet.column_dimensions["E"].width = 30

    wb.save('BK72XX_commits.xlsx')

def update_excel(reps, branch, commit, date, user, cid):
    wb = openpyxl.load_workbook("BK72XX_commits.xlsx")
    sheet = wb.active
    line_row = sheet.max_row
    line_row = line_row + 1 #在原末尾，增加新的行
    sheet.cell(row=line_row, column=EXCEL_REP_INDEX, value=reps) #写入单元格
    sheet.cell(row=line_row, column=EXCEL_BRANCH_INDEX, value=branch)
    sheet.cell(row=line_row, column=EXCEL_COMMIT_INDEX, value=commit)
    sheet.cell(row=line_row, column=EXCEL_DATE_INDEX, value=date)
    sheet.cell(row=line_row, column=EXCEL_USER_INDEX, value=user)
    sheet.cell(row=line_row, column=EXCEL_CID, value=cid)
    print(" + + + save in bk72xx_commits.xlsx + + +")
    wb.save('BK72XX_commits.xlsx')

    return True

def is_belong_range(time_create, time_start, time_end):
    tm_commit = common.get_timestamp_commit_update(time_create)
    #tm_commit = common.get_timestamp_commit(time_create)
    #tm_start = common.get_timestamp_commit(time_start)
    tm_start = common.get_timestamp_now_plus()
    tm_end = common.get_timestamp_commit(time_end)	
    if((tm_commit > tm_start) and (tm_commit < tm_end) ):
        #print(" ====== commit t  ", tm_commit, " ==== === start t  ", tm_start)
        return True
    else:
        return False

def glab_get_commits(proj, time_start, time_end):
    #获取分支
    branches = proj.branches.list()
    for branch in branches:
        commits = proj.commits.list(ref_name=branch.name, page=0, per_page=100)
        for c in commits:
            if(is_belong_range(c.created_at, time_start, time_end) == True):
                print(c.id, c.committer_name, c.created_at, c.message, '\n')
                update_excel(proj.name, branch.name, c.message, c.created_at, c.committer_name, c.id )

def open_config(file_name):
    proDir = os.path.split(os.path.realpath(__file__))[0]
    # 在当前文件路径下查找.ini文件
    configPath = os.path.join(proDir, file_name)
    print(configPath)
    conf = configparser.ConfigParser()
    # 读取.ini文件
    conf.read(configPath)
    name  = conf.get("REP","repository")
    name = name.replace(' ', '')
    repo_list = name.split(',')
    time_start = conf.get("TIME","time_start")
    time_end = conf.get("TIME", "time_end")
    return repo_list, time_start, time_end

def gitlab_start():
    gl = gitlab.Gitlab(GITLAB_URL, GITLAB_ACESS_TOKEN)
    create_excel()
    repo_list, time_start, time_end = open_config("config.ini")
    # 获取所有project的name,id
    for proj in gl.projects.list(all=True, as_list=False):
        rep_info = {"name":"res_name", "id":0}
        rep_info["name"] = proj.name
        rep_info["id"] = proj.id
        for item in repo_list:
            if (rep_info["name"] == item):
                print("item is ", item)
                glab_get_commits(proj, time_start, time_end)


if __name__ == "__main__":
    gitlab_start()
    
