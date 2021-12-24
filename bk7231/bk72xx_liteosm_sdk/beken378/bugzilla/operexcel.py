# -*- coding: utf-8 -*-
import locale
import datetime
import time
#from openpyxl import Workbook
#from openpyxl import load_workbook
import openpyxl
import string
import re
 

def fetch_excel():

    #导入需要读取的第一个Excel表格的路径
    wbin = openpyxl.load_workbook("BK72XX_commits.xlsx")
    cell_cnt = 0
    
    wbout = openpyxl.Workbook()    #创建
    #获取第一个sheet
    wsin = wbin.active
    wsout =wbout.active     
    #操作单列
    marker = []
    count = 0
    for cellin in wsin["C"]:#wsin.columns[2]: for ubantu
        count= count + 1
        if (cellin.value.find("[BUGID") != -1 or cellin.value.find("[bugid") != -1 or cellin.value.find("[BugID") != -1 \
            or cellin.value.find("[Bug") != -1 or cellin.value.find("[bug") != -1 or cellin.value.find("[BUG") != -1) \
            or cellin.value.find("[Bug id") != -1 or cellin.value.find("[bug id") != -1 or cellin.value.find("[Bug ID") != -1:
            cell_cnt = cell_cnt +1
            marker.append(count)
            wsout.cell(row = cell_cnt,column = 1,value = cellin.value)
    print(marker)
    if(len(marker) == 0):
        print("not found, it is empty")
        wbout.save("BK72XX_commits_filtered.xlsx")
        return
    print("not empty")
    # commit id
    i = 1
    j = 1
    for cellin in wsin["F"]:#wsin.columns[5]: for ubantu
        if (i == marker[j - 1]):
            wsout.cell(row = j, column = 5, value = cellin.value)
            j = j + 1
            if (j > len(marker)):
                break
        i = i + 1
    # commit author    
    i = 1
    j = 1
    for cellin in wsin["E"]:
        if (i == marker[j - 1]):
            wsout.cell(row = j, column = 6, value = cellin.value)
            j = j + 1
            if (j > len(marker)):
                break
        i = i + 1
    # commit time    
    i = 1
    j = 1
    for cellin in wsin["D"]:
        if (i == marker[j - 1]):
            wsout.cell(row = j, column = 7, value = cellin.value)
            j = j + 1
            if (j > len(marker)):
                break
        i = i + 1
    # commit repo 
    i = 1
    j = 1
    for cellin in wsin["A"]:
        if (i == marker[j - 1]):
            wsout.cell(row = j, column = 8, value = cellin.value)
            j = j + 1
            if (j > len(marker)):
                break
        i = i + 1
    # commit branch 
    i = 1
    j = 1
    for cellin in wsin["B"]:
        if (i == marker[j - 1]):
            wsout.cell(row = j, column = 9, value = cellin.value)
            j = j + 1
            if (j > len(marker)):
                break
        i = i + 1
    #保存为
    print(" + + + save in bk72xx_commits_filtered.xlsx + + +")
    wbout.save("BK72XX_commits_filtered.xlsx")


def fetch_bugzilla_id():
    cell_cnt = 0
    wbin = openpyxl.load_workbook("BK72XX_commits_filtered.xlsx")
    wsin = wbin.active
    if wsin.cell(row = 1,column = 1).value is None: 
        print("not found, it is empty")
        wbin.save("BK72XX_commits_filtered_id.xlsx")
        return
    print("not empty")
    for cellin in wsin["A"]:
        cell_cnt = cell_cnt + 1
        commit_info= cellin.value[0:20]
        bid_raw = re.findall(r"\d+",commit_info)
        wsin.cell(row = cell_cnt,column = 2,value = str(bid_raw))
        bid_medium =  str(bid_raw).lstrip('[\'00')
        wsin.cell(row = cell_cnt,column = 3,value = bid_medium)
        bid_final = bid_medium.rstrip('\']') 
        wsin.cell(row = cell_cnt,column = 4,value = bid_final)
        
    print(" + + + save in bk72xx_commits_filtered_id.xlsx + + +")
    wbin.save("BK72XX_commits_filtered_id.xlsx")    
if __name__ == "__main__":
    fetch_excel()
    fetch_bugzilla_id()
