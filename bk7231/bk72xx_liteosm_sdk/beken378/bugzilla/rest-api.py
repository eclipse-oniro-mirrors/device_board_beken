#-*- coding:utf-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook

#import findcommit
import operexcel
import pprint,requests
import json
from operexcel import fetch_excel

    # -- -- -- -- -- * -  get bugzilla version  - * -- -- -- -- -- #


#findcommit.gitlab_start()
#operexcel.fetch_excel()
#operexcel.fetch_bugzilla_id()

def main():
    response = requests.get(
        'http://192.168.0.19/bugzilla/rest/version',
    )

    pprint.pprint(response.json())

if __name__ == '__main__':
    main()



def main():

    # -- -- -- -- -- * -  log in  - * -- -- -- -- -- #
    URL = "http://192.168.0.19/bugzilla/rest/bug/2"

    API_KEY = "FXkDYcVoRLxfnQ5Rr4vZaelWAj54iPMkpkAmjVUU"
   
    headers = {"Content-type": "application/json"}
    params = {
                  "api_key": API_KEY,
              }

    resp = requests.get(URL , headers = headers, params = params)


    if resp.status_code != 200:
        print('error: ' + str(resp.status_code))
    else:
        print('Success')
        #print(json.loads(resp.text))
    # -- -- -- -- -- * -  fetch token  - * -- -- -- -- -- #
    response = requests.get(
        'http://192.168.0.19/bugzilla/rest/login?login=min.wang@bekencorp.com&password=admin1983',
    )


    pprint.pprint(response.json())

    token=dict(response.json())

    # -- -- -- -- -- * -  test if token is valid  - * -- -- -- -- -- #
    response = requests.get(
        'http://192.168.0.19/bugzilla/rest/valid_login?login=min.wang@bekencorp.com&token='+token['token'],
    )

    pprint.pprint(response.json())


    fetch_excel()
    cell_cnt = 0
    wbin = load_workbook(r'BK72XX_commits_filtered_id.xlsx')

    #获取第一个sheet
    wsin = wbin.active

    rows=[]
    for row in wsin.iter_rows():
        rows.append(row)
    #操作单列
    end = wsin.max_row
    print(end)
    print(rows)
    #for row in wsin[1:end]:
    for row in rows:
        #break condition
        if wsin.cell(row = 1,column = 1).value is None: 
            print("not found, it is empty")
            wbin.save("BK72XX_commits_filtered_id.xlsx")
            return
        #end of break condition        

        #break condition
        if(cell_cnt >= end):
            return
        #end of break
    
        commit_info= rows[cell_cnt][0].value
        commit_id= rows[cell_cnt][4].value
        commit_author = rows[cell_cnt][5].value
        commit_time = rows[cell_cnt][6].value
        commit_repo = rows[cell_cnt][7].value
        commit_branch = rows[cell_cnt][8].value
        bugid = rows[cell_cnt][3].value
        print(bugid)
        cell_cnt =cell_cnt + 1
        data = {
                  "api_key" : API_KEY,               
                  'comment':{
                    'body':"[Commit] - "+commit_id+ "\n" + "[Author] - "+commit_author+ "\n" \
                    +"[Time] - "+commit_time+"\n"+"[Repo] - "+commit_repo+"\n"+"[Branch] - " \
                    +commit_branch+ "\n" +commit_info,
                    'is_private':False,
                            }
                  #"resolution":"FIXED"
                }
    
        response = requests.put(
                 'http://192.168.0.19/bugzilla/rest/bug/'+str(bugid),
                 json = data)
        print(commit_info)
        print("put operation - put finish")
        pprint.pprint(response.json)


    #保存为
    wbin.save(r'BK72XX_commits_filtered_id.xlsx')
    print("This is end, Bye!")


if __name__ == '__main__':
    main()

 


